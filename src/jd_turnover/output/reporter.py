"""报表生成模块。输出固定3个Sheet的XLSX文件。"""

from io import BytesIO

import pandas as pd
from loguru import logger
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Sheet 顺序和名称
SHEET_RAW = "底表新"
SHEET_TOTAL = "总表-分SKU"
SHEET_ORDER = "平台采购下单表"

# 总表列宽
TOTAL_COL_WIDTHS = [18, 20, 18, 12, 26, 10, 8, 10, 10, 12, 12, 14, 12, 12, 10, 18, 18, 14]
# 采购下单表列宽
ORDER_COL_WIDTHS = [22, 16, 20, 26]


def to_excel_bytes(
    df_raw: pd.DataFrame,
    df_total: pd.DataFrame,
    df_order: pd.DataFrame,
) -> BytesIO:
    """生成符合 v5 规范的 3-Sheet XLSX 字节流。"""
    output = BytesIO()

    hdr_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    hdr_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Sheet3: 底表新
        df_raw.to_excel(writer, sheet_name=SHEET_RAW, index=False)

        # Sheet2: 总表-分SKU
        df_total.to_excel(writer, sheet_name=SHEET_TOTAL, index=False)
        ws_total = writer.sheets[SHEET_TOTAL]
        for col_idx in range(1, len(df_total.columns) + 1):
            cell = ws_total.cell(row=1, column=col_idx)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = center_align
        for i, w in enumerate(TOTAL_COL_WIDTHS, 1):
            ws_total.column_dimensions[get_column_letter(i)].width = w
        ws_total.freeze_panes = "A2"

        # Sheet1: 平台采购下单表
        df_order.to_excel(writer, sheet_name=SHEET_ORDER, index=False)
        ws_order = writer.sheets[SHEET_ORDER]
        for col_idx in range(1, len(df_order.columns) + 1):
            cell = ws_order.cell(row=1, column=col_idx)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = center_align
        for i, w in enumerate(ORDER_COL_WIDTHS, 1):
            ws_order.column_dimensions[get_column_letter(i)].width = w
        ws_order.freeze_panes = "A2"

    output.seek(0)
    logger.info(f"报表生成完成: {SHEET_RAW}({len(df_raw)}行), {SHEET_TOTAL}({len(df_total)}行), {SHEET_ORDER}({len(df_order)}行)")
    return output


def build_summary_dict(
    df_raw: pd.DataFrame,
    df_total: pd.DataFrame,
    df_order: pd.DataFrame,
) -> dict[str, pd.DataFrame]:
    """构建多 Sheet 字典（供兼容旧接口使用）。"""
    return {SHEET_RAW: df_raw, SHEET_TOTAL: df_total, SHEET_ORDER: df_order}
